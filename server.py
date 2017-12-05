#coding=utf-8
from __future__ import division
import flask
from flask import Flask
from flask import request, render_template, send_from_directory, redirect, url_for
from flask_sqlalchemy import SQLAlchemy
from flask_sse import sse
import flask_login
from flask_cors import CORS, cross_origin
import hashlib
import json
import requests
from datetime import datetime
import time
from openpyxl import Workbook, load_workbook
from concurrent.futures import ThreadPoolExecutor, as_completed

app = Flask(__name__)
CORS(app)
app.config["REDIS_URL"] = "redis://localhost:6379"
app.register_blueprint(sse, url_prefix='/stream')
app.config['SECRET_KEY'] = "123456789"

# database config
host = '106.14.63.93'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = True
app.config['SQLALCHEMY_DATABASE_URI'] = 'mysql://root:27271992@'+host+':3306/map'
db = SQLAlchemy(app)
md5 = hashlib.md5()

# login management
login_manager = flask_login.LoginManager()
login_manager.init_app(app)

# running config
baidu_api_timeout = 5 # seconds
n_threads = 100

class User(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(20), unique=True)
    password = db.Column(db.String(32))
    configs = db.relationship('Config', backref='user')

    def __init__(self, username, psd):
        self.name = username
        md5.update(psd)
        self.password = md5.hexdigest()

    def __repr__(self):
        return '<User %r %r>' % (self.name, self.id)

class Config(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(50))
    uid = db.Column(db.Integer, db.ForeignKey('user.id'))
    points = db.relationship('Point', backref='config',  passive_deletes=True)
    def __init__(self, config_name, uid):
        self.name = config_name
        self.uid = uid

    def __repr__(self):
        return '<User %r Config %r>' % (self.uid, self.name)

class Point(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    cid = db.Column(db.Integer, db.ForeignKey('config.id', ondelete='CASCADE'))
    lat = db.Column(db.Float)
    lng = db.Column(db.Float)
    ra = db.Column(db.Float)
    rb = db.Column(db.Float)
    rc = db.Column(db.Float)
    idx = db.Column(db.Integer)
    name = db.Column(db.String(50))

    def __init__(self, config_id, lat, lng, ra, rb, rc, idx, name):
        self.cid = config_id
        self.lat = lat
        self.lng = lng
        self.ra = ra
        self.rb = rb
        self.rc = rc
        self.idx = idx
        self.name = name

    def __repr__(self):
        return '<Point lat: %r lng: %r>' % (self.lat, self.lng)

class UserForAuth(flask_login.UserMixin):

    def get_id(self):
        user = User.query.filter_by(id=self.id).first()
        if user is None:
            return None
        else:
            return user.id

@login_manager.user_loader
def load_user(user_id):
    user = User.query.filter_by(id=user_id).first()
    user_auth = UserForAuth()
    user_auth.username = user.name
    user_auth.id = user.id
    return user_auth

def authenticate(name, password):
    user = User.query.filter_by(name=name).first()
    if user is None:
        return False
    else:
        md5.update(password)
        is_auth = md5.hexdigest() == user.password
        is_auth = True
        if is_auth:
            user_auth = UserForAuth()
            user_auth.username = name
            user_auth.id = user.id
            flask_login.login_user(user_auth, force=True)
            return True
        return False

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'GET':
        return render_template('login.html')

    name = request.form['userName']
    password = request.form['password']
    is_auth = authenticate(name, password)
    if not is_auth:
        return flask.jsonify({'status': 2})
    else:
        return flask.jsonify({'status': 1})

@app.route('/index', methods=['GET'])
@flask_login.login_required
def index():
    return render_template('index.html')

# use requests to use baidu api
def baidu_route(origin, destination, mode):
    pass

def baidu_direction(origin, destination, mode):
    params={
        'origin': origin,
        'destination': destination,
        'mode': mode,
        'origin_region': "上海",
        'destination_region': "上海",
        'region': '上海',
        'output': 'json',
        'ak': 'kxreayOyZO9mLxjIOotUVT8F'
    }
    baidu_url = 'http://api.map.baidu.com/direction/v1'
    try:
        ret = requests.get(url=baidu_url, params = params, timeout=baidu_api_timeout)
    except (requests.exceptions.Timeout, requests.exceptions.ConnectionError), arg:
        return -1
    response = json.loads(ret.text)
    # print response
    if response['status'] != 0 or not response['result'].has_key('routes'):
        return -1
    if mode == 'driving':
        return response['result']['routes'][0]['duration'] / 60
    else:
        return response['result']['routes'][0]['scheme'][0]['duration'] / 60

# compute the score of point
def compute(points, basic):
    start_p = str(basic['lat'])+','+str(basic['lng'])
    total = 0
    total_c = 0
    times_drive = []
    times_bus = []
    for key, p in points:
        end_p = str(p['lat']) + ',' + str(p['lng'])
        t_drive = baidu_direction(start_p, end_p, 'driving')
        t_bus = baidu_direction(start_p, end_p, 'transit')
        times_drive.append(t_drive)
        times_bus.append(t_bus)
        # print 'driver time: %d, bus time: %d ' % (t_drive, t_bus)
        if t_drive == -1 or t_bus == -1:
            print "error"
            continue
        t = (t_bus*p['ra'] + t_drive*p['rb']) / (p['ra'] + p['rb'])
        total += t * p['rc']
        total_c += p['rc']
    if total_c == 0:
        total = 0
    else:
        total = total / total_c
    return total, times_drive, times_bus


@app.route('/export', methods=['POST'])
@flask_login.login_required
def export():
    points = json.loads(request.form['points']).items()
    basic = json.loads(request.form['basic'])
    # compute the total score
    score, t_drive, t_bus = compute(points, basic)
    return flask.jsonify({'score': score}), 200

@app.route('/batch', methods=['POST'])
@flask_login.login_required
def batch():
    if 'file' not in request.files:
        flash('No file part')
        return 'no file'
    file_prefix = flask_login.current_user.username
    # file_prefix = '1'
    points = json.loads(request.form['points'])
    points = points.items()
    points.sort()
    file = request.files['file']
    name_save = file_prefix+'_'+file.filename
    file.save('./files/'+ name_save)
    wb = load_workbook('./files/'+ name_save)
    st = wb.active
    wb_result = Workbook()
    ws = wb_result.active
    ws.title = 'result'

    # header
    ws.append(('Id', 'Name', 'Lat', 'Lng', 'Ra', 'Rb', 'Rc'))
    # 参考点
    for key, p in points:
        ws.append((p['idx'], p['name'], p['lat'], p['lng'], p['ra'], p['rb'], p['rc']))
    ws.append(())

    # header
    ids = [x[0] for x in points]
    ws.append(('lat', 'lng', 'score')+tuple([str(id)+'_drive' for id in ids])
              +tuple([str(id)+'_bus' for id in ids]))
    
    n_refer_points = len(points)
    n_basic_points = st.max_row - 1
    concurrency_baidu_map = 2000
    time_window = 60 # seconds
    one_percent = n_basic_points // 100
    i = 0
    executor = ThreadPoolExecutor(max_workers=n_threads)
    # submit http requests
    futures = {}
    count = 0
    count_acc = -1
    for idx, r in enumerate(st.rows[1:]):
        basic = {'lat': r[0].value, 'lng': r[1].value}
        result_future = executor.submit(compute, points, basic)
        count += n_refer_points * 2
        count_acc += 1
        if count_acc % one_percent == 0:
            i += 1
            sse.publish({"progress": i}, type="report")
        if count > concurrency_baidu_map:
            time.sleep(time_window)
            count -= concurrency_baidu_map
        futures[result_future] = basic

    # wait for results
    timeout = n_basic_points * n_refer_points * 2 * baidu_api_timeout
    count = 0
    for f in as_completed(futures, timeout):
        basic = futures[f]
        count += 1
        try:
            score, t_drive, t_bus = f.result()
        except Exception as exc:
            print('%r generated an exception: %s' % (basic, exc))
        else:
            ws.append((basic['lat'], basic['lng'], score)+tuple(t_drive)+tuple(t_bus))
        print "%d/%d finished" % (count, n_basic_points)
    
    # save file
    file_name = file_prefix + datetime.now().strftime(' %Y-%m-%d %H:%M:%S') + '.xlsx'
    wb_result.save(filename='./files/' + file_name)
    sse.publish({"progress": 100}, type="report")
    return flask.jsonify({'file': file_name}), 200

@app.route('/save', methods=['POST'])
@flask_login.login_required
def save_config():
    refer_name = request.form['name']
    points = json.loads(request.form['points'])
    userid = flask_login.current_user.id
    config = Config(refer_name, userid)
    db.session.add(config)
    db.session.commit()
    for i, key in enumerate(points):
        p = points[key]
        point = Point(config.id, p['lat'], p['lng'],
            p['ra'], p['rb'], p['rc'], i, p['name'])
        db.session.add(point)
    db.session.commit()
    return 'success', 200

@app.route('/delete', methods=['GET'])
@flask_login.login_required
def delete_config():
    config_id = request.args["config_id"]
    config_delete = Config.query.filter_by(id=config_id).one()
    db.session.delete(config_delete)
    db.session.commit()
    return 'success', 200

@app.route('/configs', methods=['GET', 'POST'])
@flask_login.login_required
def get_config():
    userid = flask_login.current_user.id
    configs = Config.query.filter_by(uid=userid).all()
    ret = []
    for config in configs:
        ret.append({'id': config.id, 'name': config.name})
    return flask.jsonify({'status': 1, 'configs': ret})

@app.route('/points', methods=['GET'])
# @flask_login.login_required
def get_point():
    cid = request.args['cid']
    points = Point.query.filter_by(cid=cid).all()
    ret = []
    for p in points:
        ret.append({
            'lat': p.lat,
            'lng': p.lng,
            'ra': p.ra,
            'rb': p.rb,
            'rc': p.rc,
            'idx': p.idx,
            'name': p.name
        })
    return flask.jsonify({'points': ret})


@app.route('/download', methods=['GET'])
@flask_login.login_required
def download():
    if request.args.has_key('file'):
        file_name = request.args['file']
        return flask.send_from_directory('files', file_name,
    cache_timeout=1, attachment_filename=file_name, as_attachment=True), 200
    else:
        return 'no file specifid', 400

@app.route('/protected')
@flask_login.login_required
def protected():
    return 'Logged in as: ' + flask_login.current_user.username
